
import os 
import rasterio
import rasterio.plot
import pyproj
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook

header = [ "Nothing"	,"Latitude"	,"Longitude"	,"ID"	,"DEVICE",	"centroid"	,"B1"	,"B2",	"B3"	,"B4"	,"B5"	,"B6"	,"B7"	,"B8"	,"B9",	"B10"	,"B11"	,"B12",	"B13",	"B14",	"B15"	,"B16" ]
def getFieldsFromLongLat ( longitude , lattitude, geotiff ):
    # Use pyproj to :  lon lat ====>  east, north
    east,north = pyproj.transform(geotiff["lonlat"], geotiff["utm"], longitude, lattitude)    
    # east north  ==> row, col
    row, col = geotiff["src"].index(east, north) # spatial --> image coordinates
    # row, col ==> NDVI
    if ( ( (col < geotiff["src"].profile["width"] ) & (col > 0) ) & (row > 0) & (row < geotiff["src"].profile["height"]) ):
        channels = []
        for i in range (0, 16):
            channels.append  (geotiff["channel"][i][row, col]  )
        return channels
    else: 
        return ([9392])

def readGeoTiff (filename):
    #open file and set up projections
    geotiff             =   {}
    geotiff["src"]      =   rasterio.open(filename) 
    geotiff["channel"] = []
    for i in range (1,17):
        geotiff["channel"].append (  geotiff["src"].read(i) )  # read the entire array
    geotiff["utm"]      =   pyproj.Proj(geotiff["src"].crs) # Pass CRS of image from rasterio
    geotiff["lonlat"]   =   pyproj.Proj(init="epsg:4326")
    return (geotiff)

def createTabSheetInExcel_wb(wb, tabName, allLines, geotiff):
    tmpVar = []
    #opens a sheet in that workbook
    ws1 = wb.create_sheet(title=tabName) 

    ws1.append(header)
    for i in range(1,len(longitudeArray)):
        channelsAtLongLat = getFieldsFromLongLat (longitudeArray[i], latittudeArray[i], geotiff)
        if (len (channelsAtLongLat) > 2):
            tmpVar = []
            tmpVar.extend(allLines[i])
            tmpVar.extend(channelsAtLongLat)
            ws1.append ( tmpVar )

# main
# read the csv file containing the assets (long lat) and put long lat as seen.
xlsxFilename = os.getcwd() + "/data/plot_stations.xlsx"
wb = load_workbook(xlsxFilename)
print (wb.sheetnames)
ws = wb.active

latittudeArray = []
for row in [ws["B"]]:
    for value in row:
        latittudeArray.append(value.value)


longitudeArray = []
for row in [ws["C"]]:
    for value in row:
        longitudeArray.append(value.value)

# allLines will have all the lines in that file.
allLines=[]
for row in ws.rows:
    line = []
    for cell in row:
        line.append(cell.value)
    allLines.append(line)

# opens the output xlsx file
wb = Workbook()  #this is the final xlsx that contains all the data.

files = os.listdir("./data/d/allBands")

#read tiff file
for file in files:
    print(file)
    filename = os.getcwd() + "/data/d/allBands/" + file #path to tata
    geotiff = readGeoTiff (filename)
    createTabSheetInExcel_wb (wb, file, allLines, geotiff )

#save the file
dest_filename = os.getcwd() + "/data/plot_stations_sat.xlsx"
wb.save(filename = dest_filename)






"""
>>> ws2 = wb.create_sheet(title="Pi")
>>>
>>> ws2["F5"] = 3.14
>>>
>>> ws3 = wb.create_sheet(title="Data")
>>> for row in range(10, 20):
...     for col in range(27, 54):
...         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
>>> print(ws3["AA10"].value)
AA
>>> wb.save(filename = dest_filename)
"""